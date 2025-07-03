import time 
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import psycopg2
from psycopg2.extras import execute_batch
from .moysklad import MoyskladAPI
from datetime import datetime
import os
from openpyxl import Workbook
import io
import asyncio
from typing import List, Dict, Any
import logging
import uuid
from google.oauth2.service_account import Credentials
import gspread
from fastapi import HTTPException
from pathlib import Path
from fastapi.responses import JSONResponse
from datetime import datetime
from decimal import Decimal
import json
from fastapi import Response
from fastapi.responses import StreamingResponse
from urllib.parse import quote


# Настройка логгера
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Разрешает все домены (для разработки)
    allow_methods=["*"],  # Разрешает все HTTP-методы
    allow_headers=["*"],  # Разрешает все заголовки
)

# Настройки базы данных
DB_CONFIG = {
    "host": "87.228.99.200",
    "port": 5432,
    "dbname": "MS",
    "user": "louella",
    "password": "XBcMJoEO1ljb",
    "sslmode": "verify-ca",
    "sslrootcert": "/root/.postgresql/root.crt"
}

# Инициализация API МойСклад
moysklad = MoyskladAPI(token="2e61e26f0613cf33fab5f31cf105302aa2d607c3")

class DateRange(BaseModel):
    start_date: str
    end_date: str

class BatchProcessResponse(BaseModel):
    task_id: str
    status: str
    message: str

# Глобальный словарь для хранения статусов задач
tasks_status = {}

def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)

def init_db():
    """Инициализация базы данных - создание таблиц если они не существуют"""
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Проверяем существование таблицы demands
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'demands'
            )
        """)
        demands_exists = cur.fetchone()[0]
        
        if not demands_exists:
            cur.execute("""
                CREATE TABLE demands (
                    id VARCHAR(255) PRIMARY KEY,
                    number VARCHAR(50),
                    date TIMESTAMP,
                    counterparty VARCHAR(255),
                    store VARCHAR(255),
                    project VARCHAR(255),
                    sales_channel VARCHAR(255),
                    amount NUMERIC(15, 2),
                    cost_price NUMERIC(15, 2),
                    overhead NUMERIC(15, 2),
                    profit NUMERIC(15, 2),
                    promo_period VARCHAR(255),
                    delivery_amount NUMERIC(15, 2),
                    admin_data NUMERIC(15, 2),
                    gdeslon NUMERIC(15, 2),
                    cityads NUMERIC(15, 2),
                    ozon NUMERIC(15, 2),
                    ozon_fbs NUMERIC(15, 2),
                    yamarket_fbs NUMERIC(15, 2),
                    yamarket_dbs NUMERIC(15, 2),
                    yandex_direct NUMERIC(15, 2),
                    price_ru NUMERIC(15, 2),
                    wildberries NUMERIC(15, 2),
                    gis2 NUMERIC(15, 2),
                    seo NUMERIC(15, 2),
                    programmatic NUMERIC(15, 2),
                    avito NUMERIC(15, 2),
                    multiorders NUMERIC(15, 2),
                    estimated_discount NUMERIC(15, 2),
                    status VARCHAR(100),
                    comment VARCHAR(255)
                )
            """)
            logger.info("Таблица demands успешно создана")
        
        # Проверяем существование таблицы demand_positions
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'demand_positions'
            )
        """)
        positions_exists = cur.fetchone()[0]
        
        if not positions_exists:
            cur.execute("""
                CREATE TABLE demand_positions (
                    id VARCHAR(255) PRIMARY KEY,
                    demand_id VARCHAR(255) REFERENCES demands(id),
                    demand_number VARCHAR(50),
                    date TIMESTAMP,
                    counterparty VARCHAR(255),
                    store VARCHAR(255),
                    project VARCHAR(255),
                    sales_channel VARCHAR(255),
                    product_name VARCHAR(255),
                    quantity NUMERIC(15, 3),
                    price NUMERIC(15, 2),
                    amount NUMERIC(15, 2),
                    cost_price NUMERIC(15, 2),
                    article VARCHAR(100),
                    code VARCHAR(100),
                    overhead NUMERIC(15, 2),
                    profit NUMERIC(15, 2),
                    promo_period VARCHAR(255),
                    delivery_amount NUMERIC(15, 2),
                    admin_data NUMERIC(15, 2),
                    gdeslon NUMERIC(15, 2),
                    cityads NUMERIC(15, 2),
                    ozon NUMERIC(15, 2),
                    ozon_fbs NUMERIC(15, 2),
                    yamarket_fbs NUMERIC(15, 2),
                    yamarket_dbs NUMERIC(15, 2),
                    yandex_direct NUMERIC(15, 2),
                    price_ru NUMERIC(15, 2),
                    wildberries NUMERIC(15, 2),
                    gis2 NUMERIC(15, 2),
                    seo NUMERIC(15, 2),
                    programmatic NUMERIC(15, 2),
                    avito NUMERIC(15, 2),
                    multiorders NUMERIC(15, 2),
                    estimated_discount NUMERIC(15, 2)
                )
            """)
            logger.info("Таблица demand_positions успешно создана")
        
        conn.commit()
        
    except Exception as e:
        logger.error(f"Ошибка при инициализации базы данных: {str(e)}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()

@app.on_event("startup")
async def startup_event():
    """Действия при старте приложения"""
    init_db()
    logger.info("Приложение запущено, база данных инициализирована")

async def process_demands_batch(demands: List[Dict[str, Any]], task_id: str):
    """Асинхронная обработка пакета отгрузок с улучшенным логированием"""
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        batch_size = 50
        saved_count = 0
        total_count = len(demands)
        
        logger.info(f"Начало обработки {total_count} отгрузок")
        
        demands_batch = []
        positions_batch = []
        
        for idx, demand in enumerate(demands, 1):
            try:
                # Подготовка данных отгрузки
                demand_values = prepare_demand_data(demand)
                demands_batch.append(demand_values)
                
                # Подготовка данных позиций
                positions = demand.get("positions", [])
                for position in positions:
                    position_values = prepare_position_data(demand, position)
                    positions_batch.append(position_values)
                
                if len(demands_batch) >= batch_size:
                    # Вставляем отгрузки
                    inserted_demands = await insert_demands_batch(cur, demands_batch)
                    saved_count += inserted_demands
                    
                    # Вставляем позиции
                    await insert_positions_batch(cur, positions_batch)
                    
                    demands_batch = []
                    positions_batch = []
                    
                    # Обновляем статус задачи
                    if idx % 100 == 0:
                        logger.info(f"Обработано {idx}/{total_count} записей")
                        tasks_status[task_id] = {
                            "status": "processing",
                            "progress": f"{saved_count}/{total_count}",
                            "message": f"Обработано {idx} из {total_count}"
                        }
                    
                    time.sleep(0.5)
            
            except Exception as e:
                logger.error(f"Ошибка при обработке отгрузки {demand.get('id')}: {str(e)}")
                continue
        
        # Вставляем оставшиеся записи
        if demands_batch:
            saved_count += await insert_demands_batch(cur, demands_batch)
        if positions_batch:
            await insert_positions_batch(cur, positions_batch)
        
        conn.commit()
        logger.info(f"Успешно сохранено {saved_count} из {total_count} записей")
        
        tasks_status[task_id] = {
            "status": "completed",
            "progress": f"{saved_count}/{total_count}",
            "message": f"Успешно сохранено {saved_count} из {total_count} записей"
        }
        
    except Exception as e:
        logger.error(f"Критическая ошибка при обработке пакета: {str(e)}")
        if conn:
            conn.rollback()
        tasks_status[task_id] = {
            "status": "failed",
            "progress": f"{saved_count}/{total_count}",
            "message": f"Ошибка: {str(e)}"
        }
    finally:
        if conn:
            conn.close()

async def insert_demands_batch(cur, batch_values: List[Dict[str, Any]]) -> int:
    """Массовая вставка пакета данных отгрузок"""
    try:
        query = """
            INSERT INTO demands (
                id, number, date, counterparty, store, project, sales_channel, 
                amount, cost_price, overhead, profit, promo_period, delivery_amount,
                admin_data, gdeslon, cityads, ozon, ozon_fbs, yamarket_fbs,
                yamarket_dbs, yandex_direct, price_ru, wildberries, gis2, seo,
                programmatic, avito, multiorders, estimated_discount, status, comment
            ) VALUES (
                %(id)s, %(number)s, %(date)s, %(counterparty)s, %(store)s, %(project)s, %(sales_channel)s,
                %(amount)s, %(cost_price)s, %(overhead)s, %(profit)s, %(promo_period)s, %(delivery_amount)s,
                %(admin_data)s, %(gdeslon)s, %(cityads)s, %(ozon)s, %(ozon_fbs)s, %(yamarket_fbs)s,
                %(yamarket_dbs)s, %(yandex_direct)s, %(price_ru)s, %(wildberries)s, %(gis2)s, %(seo)s,
                %(programmatic)s, %(avito)s, %(multiorders)s, %(estimated_discount)s, %(status)s, %(comment)s
            )
            ON CONFLICT (id) DO UPDATE SET
                number = EXCLUDED.number,
                date = EXCLUDED.date,
                counterparty = EXCLUDED.counterparty,
                store = EXCLUDED.store,
                project = EXCLUDED.project,
                sales_channel = EXCLUDED.sales_channel,
                amount = EXCLUDED.amount,
                cost_price = EXCLUDED.cost_price,
                overhead = EXCLUDED.overhead,
                profit = EXCLUDED.profit,
                promo_period = EXCLUDED.promo_period,
                delivery_amount = EXCLUDED.delivery_amount,
                admin_data = EXCLUDED.admin_data,
                gdeslon = EXCLUDED.gdeslon,
                cityads = EXCLUDED.cityads,
                ozon = EXCLUDED.ozon,
                ozon_fbs = EXCLUDED.ozon_fbs,
                yamarket_fbs = EXCLUDED.yamarket_fbs,
                yamarket_dbs = EXCLUDED.yamarket_dbs,
                yandex_direct = EXCLUDED.yandex_direct,
                price_ru = EXCLUDED.price_ru,
                wildberries = EXCLUDED.wildberries,
                gis2 = EXCLUDED.gis2,
                seo = EXCLUDED.seo,
                programmatic = EXCLUDED.programmatic,
                avito = EXCLUDED.avito,
                multiorders = EXCLUDED.multiorders,
                estimated_discount = EXCLUDED.estimated_discount,
                status = EXCLUDED.status,
                comment = EXCLUDED.comment
        """
        
        execute_batch(cur, query, batch_values)
        return len(batch_values)
    
    except Exception as e:
        logger.error(f"Ошибка при массовой вставке отгрузок: {str(e)}")
        return 0

async def insert_positions_batch(cur, batch_values: List[Dict[str, Any]]) -> int:
    """Массовая вставка пакета данных позиций"""
    try:
        query = """
            INSERT INTO demand_positions (
                id, demand_id, demand_number, date, counterparty, store, project, sales_channel,
                product_name, quantity, price, amount, cost_price, article, code, overhead, profit,
                promo_period, delivery_amount, admin_data, gdeslon, cityads, ozon, ozon_fbs,
                yamarket_fbs, yamarket_dbs, yandex_direct, price_ru, wildberries, gis2, seo,
                programmatic, avito, multiorders, estimated_discount
            ) VALUES (
                %(id)s, %(demand_id)s, %(demand_number)s, %(date)s, %(counterparty)s, %(store)s, %(project)s, %(sales_channel)s,
                %(product_name)s, %(quantity)s, %(price)s, %(amount)s, %(cost_price)s, %(article)s, %(code)s, %(overhead)s, %(profit)s,
                %(promo_period)s, %(delivery_amount)s, %(admin_data)s, %(gdeslon)s, %(cityads)s, %(ozon)s, %(ozon_fbs)s,
                %(yamarket_fbs)s, %(yamarket_dbs)s, %(yandex_direct)s, %(price_ru)s, %(wildberries)s, %(gis2)s, %(seo)s,
                %(programmatic)s, %(avito)s, %(multiorders)s, %(estimated_discount)s
            )
            ON CONFLICT (id) DO UPDATE SET
                demand_id = EXCLUDED.demand_id,
                demand_number = EXCLUDED.demand_number,
                date = EXCLUDED.date,
                counterparty = EXCLUDED.counterparty,
                store = EXCLUDED.store,
                project = EXCLUDED.project,
                sales_channel = EXCLUDED.sales_channel,
                product_name = EXCLUDED.product_name,
                quantity = EXCLUDED.quantity,
                price = EXCLUDED.price,
                amount = EXCLUDED.amount,
                cost_price = EXCLUDED.cost_price,
                article = EXCLUDED.article,
                code = EXCLUDED.code,
                overhead = EXCLUDED.overhead,
                profit = EXCLUDED.profit,
                promo_period = EXCLUDED.promo_period,
                delivery_amount = EXCLUDED.delivery_amount,
                admin_data = EXCLUDED.admin_data,
                gdeslon = EXCLUDED.gdeslon,
                cityads = EXCLUDED.cityads,
                ozon = EXCLUDED.ozon,
                ozon_fbs = EXCLUDED.ozon_fbs,
                yamarket_fbs = EXCLUDED.yamarket_fbs,
                yamarket_dbs = EXCLUDED.yamarket_dbs,
                yandex_direct = EXCLUDED.yandex_direct,
                price_ru = EXCLUDED.price_ru,
                wildberries = EXCLUDED.wildberries,
                gis2 = EXCLUDED.gis2,
                seo = EXCLUDED.seo,
                programmatic = EXCLUDED.programmatic,
                avito = EXCLUDED.avito,
                multiorders = EXCLUDED.multiorders,
                estimated_discount = EXCLUDED.estimated_discount
        """
        
        execute_batch(cur, query, batch_values)
        return len(batch_values)
    
    except Exception as e:
        logger.error(f"Ошибка при массовой вставке позиций: {str(e)}")
        return 0

def prepare_demand_data(demand: Dict[str, Any]) -> Dict[str, Any]:
    """Подготовка данных отгрузки для вставки в БД"""
    demand_id = str(demand.get("id", ""))
    attributes = demand.get("attributes", [])
    
    # Обработка накладных расходов (overhead)
    overhead_data = demand.get("overhead", {})
    overhead_sum = float(overhead_data.get("sum", 0)) / 100
    
    # Получаем себестоимость
    cost_price = moysklad.get_demand_cost_price(demand_id)
    demand_sum = float(demand.get("sum", 0)) / 100
    profit = demand_sum - cost_price - overhead_sum
    
    # Основные данные
    values = {
        "id": demand_id[:255],
        "number": str(demand.get("name", ""))[:50],
        "date": demand.get("moment", ""),
        "counterparty": str(demand.get("agent", {}).get("name", ""))[:255],
        "store": str(demand.get("store", {}).get("name", ""))[:255],
        "project": str(demand.get("project", {}).get("name", "Без проекта"))[:255],
        "sales_channel": str(demand.get("salesChannel", {}).get("name", "Без канала"))[:255],
        "amount": demand_sum,
        "cost_price": cost_price,
        "overhead": overhead_sum,
        "profit": profit,
        "status": str(demand.get("state", {}).get("name", ""))[:100],
        "comment": str(demand.get("description", ""))[:255],
        "promo_period": "",
        "delivery_amount": 0,
        "admin_data": 0,
        "gdeslon": 0,
        "cityads": 0,
        "ozon": 0,
        "ozon_fbs": 0,
        "yamarket_fbs": 0,
        "yamarket_dbs": 0,
        "yandex_direct": 0,
        "price_ru": 0,
        "wildberries": 0,
        "gis2": 0,
        "seo": 0,
        "programmatic": 0,
        "avito": 0,
        "multiorders": 0,
        "estimated_discount": 0
    }

    # Обработка атрибутов
    attr_fields = {
        "promo_period": ("Акционный период", ""),
        "delivery_amount": ("Сумма доставки", 0),
        "admin_data": ("Адмидат", 0),
        "gdeslon": ("ГдеСлон", 0),
        "cityads": ("CityAds", 0),
        "ozon": ("Ozon", 0),
        "ozon_fbs": ("Ozon FBS", 0),
        "yamarket_fbs": ("Яндекс Маркет FBS", 0),
        "yamarket_dbs": ("Яндекс Маркет DBS", 0),
        "yandex_direct": ("Яндекс Директ", 0),
        "price_ru": ("Price ru", 0),
        "wildberries": ("Wildberries", 0),
        "gis2": ("2Gis", 0),
        "seo": ("SEO", 0),
        "programmatic": ("Программатик", 0),
        "avito": ("Авито", 0),
        "multiorders": ("Мультиканальные заказы", 0),
        "estimated_discount": ("Примерная скидка", 0)
    }

    for field, (attr_name, default) in attr_fields.items():
        if field.endswith("_amount") or field == "estimated_discount":
            try:
                values[field] = float(get_attr_value(attributes, attr_name, default))
            except (ValueError, TypeError):
                values[field] = 0.0
        else:
            values[field] = str(get_attr_value(attributes, attr_name, default))[:255]
    
    return values

def prepare_position_data(demand: Dict[str, Any], position: Dict[str, Any]) -> Dict[str, Any]:
    """Подготовка данных позиции для вставки в БД"""
    position_id = str(position.get("id", ""))
    demand_id = str(demand.get("id", ""))
    attributes = demand.get("attributes", [])
    
    # Получаем себестоимость позиции (уже в рублях)
    cost_price = position.get("cost_price", 0.0)
    
    # Количество и цена
    quantity = float(position.get("quantity", 0))
    price = float(position.get("price", 0)) / 100
    amount = quantity * price
    
    # Накладные расходы (overhead) из данных отгрузки
    overhead_data = demand.get("overhead", {})
    overhead_sum = (float(overhead_data.get("sum", 0)) / 100) if overhead_data else 0
    
    # Расчет доли накладных расходов для позиции
    demand_sum = float(demand.get("sum", 0)) / 100
    overhead_share = overhead_sum * (amount / demand_sum) if demand_sum > 0 else 0
    
    # Основные данные
    values = {
        "id": position_id[:255],
        "demand_id": demand_id[:255],
        "demand_number": str(demand.get("name", ""))[:50],
        "date": demand.get("moment", ""),
        "counterparty": str(demand.get("agent", {}).get("name", ""))[:255],
        "store": str(demand.get("store", {}).get("name", ""))[:255],
        "project": str(demand.get("project", {}).get("name", "Без проекта"))[:255],
        "sales_channel": str(demand.get("salesChannel", {}).get("name", "Без канала"))[:255],
        "product_name": str(position.get("product_name", ""))[:255],
        "quantity": quantity,
        "price": price,
        "amount": amount,
        "cost_price": cost_price,  # Себестоимость позиции
        "article": str(position.get("article", ""))[:100],
        "code": str(position.get("code", ""))[:100],
        "overhead": overhead_share,
        "profit": amount - cost_price - overhead_share,
        "promo_period": "",
        "delivery_amount": 0,
        "admin_data": 0,
        "gdeslon": 0,
        "cityads": 0,
        "ozon": 0,
        "ozon_fbs": 0,
        "yamarket_fbs": 0,
        "yamarket_dbs": 0,
        "yandex_direct": 0,
        "price_ru": 0,
        "wildberries": 0,
        "gis2": 0,
        "seo": 0,
        "programmatic": 0,
        "avito": 0,
        "multiorders": 0,
        "estimated_discount": 0
    }

    # Обработка атрибутов
    attr_fields = {
        "promo_period": ("Акционный период", ""),
        "delivery_amount": ("Сумма доставки", 0),
        "admin_data": ("Адмидат", 0),
        "gdeslon": ("ГдеСлон", 0),
        "cityads": ("CityAds", 0),
        "ozon": ("Ozon", 0),
        "ozon_fbs": ("Ozon FBS", 0),
        "yamarket_fbs": ("Яндекс Маркет FBS", 0),
        "yamarket_dbs": ("Яндекс Маркет DBS", 0),
        "yandex_direct": ("Яндекс Директ", 0),
        "price_ru": ("Price ru", 0),
        "wildberries": ("Wildberries", 0),
        "gis2": ("2Gis", 0),
        "seo": ("SEO", 0),
        "programmatic": ("Программатик", 0),
        "avito": ("Авито", 0),
        "multiorders": ("Мультиканальные заказы", 0),
        "estimated_discount": ("Примерная скидка", 0)
    }

    for field, (attr_name, default) in attr_fields.items():
        if field.endswith("_amount") or field == "estimated_discount":
            try:
                values[field] = float(get_attr_value(attributes, attr_name, default))
            except (ValueError, TypeError):
                values[field] = 0.0
        else:
            values[field] = str(get_attr_value(attributes, attr_name, default))[:255]
    
    return values

def get_attr_value(attrs, attr_name, default=""):
    """Безопасное извлечение атрибутов"""
    if not attrs:
        return default
    for attr in attrs:
        if attr.get("name") == attr_name:
            value = attr.get("value")
            if isinstance(value, dict):
                return value.get("name", str(value))
            return str(value) if value is not None else default
    return default

@app.post("/api/save-to-db")
async def save_to_db(date_range: DateRange, background_tasks: BackgroundTasks):
    """Запуск фоновой задачи для обработки данных"""
    try:
        task_id = str(uuid.uuid4())
        tasks_status[task_id] = {
            "status": "pending",
            "progress": "0/0",
            "message": "Задача поставлена в очередь"
        }
        
        # Запускаем фоновую задачу
        background_tasks.add_task(process_data_task, date_range, task_id)
        
        return {
            "task_id": task_id,
            "status": "started",
            "message": "Обработка данных начата. Используйте task_id для проверки статуса."
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

async def process_data_task(date_range: DateRange, task_id: str):
    """Фоновая задача для обработки данных"""
    try:
        tasks_status[task_id] = {
            "status": "fetching",
            "progress": "0/0",
            "message": "Получение данных из МойСклад..."
        }
        
        # Получаем данные из МойСклад
        demands = moysklad.get_demands(date_range.start_date, date_range.end_date)
        
        if not demands:
            tasks_status[task_id] = {
                "status": "completed",
                "progress": "0/0",
                "message": "Нет данных для сохранения"
            }
            return
        
        tasks_status[task_id] = {
            "status": "processing",
            "progress": f"0/{len(demands)}",
            "message": "Начало обработки данных..."
        }
        
        # Обрабатываем данные пакетами
        await process_demands_batch(demands, task_id)
    
    except Exception as e:
        logger.error(f"Ошибка в фоновой задаче: {str(e)}")
        tasks_status[task_id] = {
            "status": "failed",
            "progress": "0/0",
            "message": f"Ошибка: {str(e)}"
        }

@app.get("/api/task-status/{task_id}")
async def get_task_status(task_id: str):
    """Проверка статуса задачи"""
    status = tasks_status.get(task_id, {
        "status": "not_found",
        "progress": "0/0",
        "message": "Задача не найдена"
    })
    return {"task_id": task_id, **status}

@app.post("/api/export/excel")
async def export_excel(date_range: DateRange):
    conn = None
    try:
        logger.info(f"Starting Excel export for {date_range.start_date} to {date_range.end_date}")
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Создаем Excel файл
        wb = Workbook()
        
        # Лист с отгрузками
        await create_demands_sheet(wb, cur, date_range)
        
        # Лист с товарами
        await create_positions_sheet(wb, cur, date_range)
        
        # Лист со сводным отчетом по товарам
        await create_summary_sheet(wb, cur, date_range)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Формируем имя файла без недопустимых символов
        safe_start = date_range.start_date.replace(':', '_').replace(' ', '_')
        safe_end = date_range.end_date.replace(':', '_').replace(' ', '_')
        filename = f"Отчет_{safe_start}_по_{safe_end}.xlsx"
        
        # Возвращаем файл как StreamingResponse
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={quote(filename)}"
            }
        )
        
    except Exception as e:
        logger.error(f"Error during Excel export: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка формирования отчета: {str(e)}")
    finally:
        if conn:
            conn.close()

async def create_demands_sheet(wb, cur, date_range):
    """Создает лист с отгрузками"""
    cur.execute("""
        SELECT 
            number, date, counterparty, store, project, sales_channel,
            amount, cost_price, overhead, profit, promo_period, delivery_amount,
            admin_data, gdeslon, cityads, ozon, ozon_fbs, yamarket_fbs,
            yamarket_dbs, yandex_direct, price_ru, wildberries, gis2, seo,
            programmatic, avito, multiorders, estimated_discount
        FROM demands
        WHERE date BETWEEN %s AND %s
        ORDER BY date DESC
    """, (date_range.start_date, date_range.end_date))
    
    rows = cur.fetchall()
    
    ws = wb.active
    ws.title = "Отчет по отгрузкам"
    
    # Заголовки столбцов
    headers = [
        "Номер отгрузки", "Дата", "Контрагент", "Склад", "Проект", "Канал продаж",
        "Сумма", "Себестоимость", "Накладные расходы", "Прибыль", "Акционный период",
        "Сумма доставки", "Адмидат", "ГдеСлон", "CityAds", "Ozon", "Ozon FBS",
        "Яндекс Маркет FBS", "Яндекс Маркет DBS", "Яндекс Директ", "Price ru",
        "Wildberries", "2Gis", "SEO", "Программатик", "Авито", "Мультиканальные заказы",
        "Примерная скидка"
    ]
    
    apply_sheet_styling(ws, headers, rows, numeric_columns=[7, 8, 9, 10, 12] + list(range(13, 29)), 
                        profit_column=10, sheet_type="demands")

async def create_positions_sheet(wb, cur, date_range):
    """Создает лист с товарами, сгруппированными по отгрузкам с себестоимостью"""
    cur.execute("""
        SELECT 
            d.number as demand_number, 
            d.date, 
            d.counterparty, 
            d.store, 
            d.project, 
            d.sales_channel,
            dp.product_name, 
            dp.quantity, 
            dp.price, 
            dp.amount, 
            dp.cost_price,
            dp.article, 
            dp.code,
            dp.overhead, 
            dp.profit, 
            d.promo_period, 
            d.delivery_amount, 
            d.admin_data, 
            d.gdeslon,
            d.cityads, 
            d.ozon, 
            d.ozon_fbs, 
            d.yamarket_fbs, 
            d.yamarket_dbs, 
            d.yandex_direct,
            d.price_ru, 
            d.wildberries, 
            d.gis2, 
            d.seo, 
            d.programmatic, 
            d.avito, 
            d.multiorders,
            d.estimated_discount,
            d.cost_price as total_cost_price
        FROM demand_positions dp
        JOIN demands d ON dp.demand_id = d.id
        WHERE d.date BETWEEN %s AND %s
        ORDER BY d.number, d.date DESC
    """, (date_range.start_date, date_range.end_date))
    
    rows = cur.fetchall()
    
    ws = wb.create_sheet("Отчет по товарам")
    
    # Заголовки столбцов
    headers = [
        "Номер отгрузки", "Дата", "Контрагент", "Склад", "Проект", "Канал продаж",
        "Товар", "Количество", "Цена", "Сумма", "Себестоимость", "Артикул", "Код",
        "Накладные расходы", "Прибыль", "Акционный период", "Сумма доставки", "Адмидат",
        "ГдеСлон", "CityAds", "Ozon", "Ozon FBS", "Яндекс Маркет FBS", "Яндекс Маркет DBS",
        "Яндекс Директ", "Price ru", "Wildberries", "2Gis", "SEO", "Программатик", "Авито",
        "Мультиканальные заказы", "Примерная скидка"
    ]
    
    # Добавляем заголовки
    ws.append(headers)
    
    # Стили для Excel
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Стиль заголовков
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Применяем стили к заголовкам
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        # Автоподбор ширины столбца
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = max(15, len(headers[col-1]) * 1.1)
    
    # Основной стиль для данных
    data_font = Font(name='Calibri', size=10)
    money_format = '#,##0.00'
    date_format = 'DD.MM.YYYY HH:MM:SS'
    
    # Группировка по отгрузкам
    current_demand = None
    row_num = 2
    
    for row in rows:
        demand_number = row[0]
        
        # Новая отгрузка - добавляем строку с итогами
        if demand_number != current_demand:
            current_demand = demand_number
            
            # Формируем строку с итогами по отгрузке
            ws.append([
                demand_number,     # Номер
                row[1],           # Дата
                row[2],           # Контрагент
                row[3],           # Склад
                row[4],           # Проект
                row[5],           # Канал
                "Итого по отгрузке:", # Товар
                "",               # Количество
                "",               # Цена
                row[9],           # Сумма
                row[33],          # Общая себестоимость
                "",              # Артикул
                "",              # Код
                row[13],         # Накладные расходы
                row[14],         # Прибыль
                row[15],         # Акционный период
                row[16],         # Сумма доставки
                row[17],         # Адмидат
                row[18],         # ГдеСлон
                row[19],         # CityAds
                row[20],         # Ozon
                row[21],         # Ozon FBS
                row[22],         # Яндекс Маркет FBS
                row[23],         # Яндекс Маркет DBS
                row[24],         # Яндекс Директ
                row[25],         # Price ru
                row[26],         # Wildberries
                row[27],         # 2Gis
                row[28],         # SEO
                row[29],         # Программатик
                row[30],         # Авито
                row[31],         # Мультиканальные заказы
                row[32]          # Примерная скидка
            ])
            
            # Применяем стили к строке с итогами
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = Font(name='Calibri', bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.border = thin_border
                
                # Форматирование числовых полей
                if col in [10, 11, 14, 15, 17] + list(range(18, 34)):
                    try:
                        cell.number_format = money_format
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except:
                        pass
                
                # Особое форматирование для "Итого по отгрузке:"
                if col == 7:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            row_num += 1
        
        # Добавляем строку с товаром
        ws.append([
            "",              # Номер
            "",              # Дата
            "",              # Контрагент
            "",              # Склад
            "",              # Проект
            "",              # Канал
            row[6],          # Товар
            row[7],          # Количество
            row[8],          # Цена
            row[9],          # Сумма
            row[10],         # Себестоимость позиции
            row[11],         # Артикул
            row[12],         # Код
            "",              # Накладные расходы
            "",              # Прибыль
            "",              # Акционный период
            "",              # Сумма доставки
            "",              # Адмидат
            "",              # ГдеСлон
            "",              # CityAds
            "",              # Ozon
            "",              # Ozon FBS
            "",              # Яндекс Маркет FBS
            "",              # Яндекс Маркет DBS
            "",              # Яндекс Директ
            "",              # Price ru
            "",              # Wildberries
            "",              # 2Gis
            "",              # SEO
            "",              # Программатик
            "",              # Авито
            "",              # Мультиканальные заказы
            ""               # Примерная скидка
        ])
        
        # Применяем стили к строке с товаром
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = data_font
            cell.border = thin_border
            
            # Форматирование числовых полей
            if col in [8, 9, 10, 11]:  # Количество, Цена, Сумма, Себестоимость
                try:
                    cell.number_format = money_format
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                except:
                    pass
            
            # Форматирование даты
            elif col == 2:
                cell.number_format = date_format
        
        row_num += 1
    
    # Добавляем автофильтр и замораживаем заголовки
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    
    # Добавляем итоговую строку
    total_row = row_num + 1
    ws.append([""] * len(headers))
    
    # Формируем итоговую строку
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        
        # Суммы для числовых столбцов
        if col in [10, 11, 14, 15, 17] + list(range(18, 34)):
            column_letter = get_column_letter(col)
            formula = f"SUM({column_letter}2:{column_letter}{row_num})"
            cell.value = f"=ROUND({formula}, 2)"
            cell.number_format = money_format
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        elif col == 1:
            cell.value = "Общий итог:"
            cell.alignment = Alignment(horizontal='right', vertical='center')

async def create_summary_sheet(wb, cur, date_range):
    """Создает лист со сводным отчетом по товарам"""
    cur.execute("""
        SELECT 
            dp.product_name,
            dp.article,
            dp.code,
            SUM(dp.quantity) as total_quantity,
            d.store,
            d.project,
            d.sales_channel,
            AVG(dp.price) as avg_price,
            SUM(d.delivery_amount * (dp.amount / NULLIF(d.amount, 0))) as delivery_share,
            SUM(dp.amount) as total_amount,
            SUM(dp.cost_price) as total_cost_price,
            SUM(d.overhead * (dp.amount / NULLIF(d.amount, 0))) as overhead_share,
            SUM(dp.profit) as total_profit,
            CASE 
                WHEN SUM(dp.amount) = 0 THEN 0 
                ELSE (SUM(dp.amount) - SUM(dp.cost_price) - SUM(d.overhead * (dp.amount / NULLIF(d.amount, 0)))) / SUM(dp.amount) * 100 
            END as margin
        FROM demand_positions dp
        JOIN demands d ON dp.demand_id = d.id
        WHERE d.date BETWEEN %s AND %s
        GROUP BY dp.product_name, dp.article, dp.code, d.store, d.project, d.sales_channel
        ORDER BY total_amount DESC
    """, (date_range.start_date, date_range.end_date))
    
    rows = cur.fetchall()
    
    ws = wb.create_sheet("Сводный отчет по товарам")
    
    # Заголовки столбцов
    headers = [
        "Товар", "Артикул", "Код", "Общее количество", "Склад", "Проект", 
        "Канал продаж", "Средняя цена", "Сумма оплачиваемой доставки", 
        "Общая сумма", "Себестоимость товара", "Сумма накладных расходов", 
        "Общая прибыль", "Маржинальность"
    ]
    
    # Стили для Excel
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Стиль заголовков
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Добавляем заголовки
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        # Автоподбор ширины столбца
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = max(15, len(headers[col-1]) * 1.1)
    
    # Основной стиль для данных
    data_font = Font(name='Calibri', size=10)
    money_format = '#,##0.00'
    percent_format = '0.00%'
    
    # Добавляем данные
    for row in rows:
        ws.append(row)
    
    # Форматируем числовые столбцы
    numeric_columns = [4, 8, 9, 10, 11, 12, 13]  # Номера столбцов с числами
    percent_column = 14  # Маржинальность
    
    for row_idx in range(2, len(rows) + 2):
        for col_idx in numeric_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = money_format
            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Форматируем столбец с маржинальностью как процент
        cell = ws.cell(row=row_idx, column=percent_column)
        cell.number_format = percent_format
        cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Подсветка отрицательной маржинальности
    for row_idx in range(2, len(rows) + 2):
        cell = ws.cell(row=row_idx, column=percent_column)
        if cell.value is not None and cell.value < 0:
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Добавляем итоговую строку
    total_row = len(rows) + 2
    ws.append([""] * len(headers))
    
    # Формируем итоговую строку
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        
        # Суммы для числовых столбцов
        if col in numeric_columns + [percent_column]:
            column_letter = get_column_letter(col)
            if col == percent_column:
                # Для маржинальности считаем средневзвешенное значение
                formula = f"=SUM(J2:J{total_row-1})"
                formula_cost = f"=SUM(K2:K{total_row-1})"
                formula_overhead = f"=SUM(L2:L{total_row-1})"
                cell.value = f"=IF({formula}=0, 0, ({formula}-{formula_cost}-{formula_overhead})/{formula})"
            else:
                formula = f"SUM({column_letter}2:{column_letter}{total_row-1})"
                cell.value = f"=ROUND({formula}, 2)"
            
            cell.number_format = money_format if col != percent_column else percent_format
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        elif col == 1:
            cell.value = "Общий итог:"
            cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Добавляем автофильтр и замораживаем заголовки
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

def apply_sheet_styling(ws, headers, rows, numeric_columns, profit_column, sheet_type):
    """Применяет стили к листу Excel"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Шрифты
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    cell_font = Font(name='Calibri', size=11)
    
    # Выравнивание
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # Границы
    thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
    
    # Заливка
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    money_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Добавляем заголовки
    ws.append(headers)
    
    # Форматируем заголовки
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
        
        # Автоподбор ширины столбца
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = max(15, len(headers[col-1]) * 1.2)
    
    # Добавляем данные и форматируем их
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            
            # Форматирование чисел и дат
            if col_idx in numeric_columns:  # Все числовые столбцы
                try:
                    num_value = float(value) if value not in [None, ''] else 0.0
                    cell.value = num_value
                    
                    # Форматирование в зависимости от типа данных
                    if sheet_type == "positions" and col_idx in [8, 9]:  # Количество и цена
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '#,##0.00'
                    
                    cell.alignment = right_alignment
                    
                    # Проверяем отрицательную прибыль
                    if col_idx == profit_column and num_value < 0:
                        cell.fill = negative_fill
                    elif row_idx % 2 == 0:  # Зебра для читаемости
                        cell.fill = money_fill
                except (ValueError, TypeError):
                    cell.alignment = left_alignment
            elif col_idx == 2:  # Столбец с датой
                cell.number_format = 'DD.MM.YYYY'
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
    
    # Замораживаем заголовки
    ws.freeze_panes = 'A2'
    
    # Добавляем автофильтр
    ws.auto_filter.ref = ws.dimensions
    
    # Добавляем итоговую строку
    last_row = len(rows) + 1
    ws.append([""] * len(headers))
    total_row = last_row + 1
    
    # Форматируем итоговую строку
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        
        # Суммы для числовых столбцов
        if col in numeric_columns:
            start_col = get_column_letter(col)
            formula = f"SUM({start_col}2:{start_col}{last_row})"
            cell.value = f"=ROUND({formula}, 2)"
            cell.number_format = '#,##0.00'
            cell.alignment = right_alignment
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        elif col == 1:
            cell.value = "Итого:"
            cell.alignment = right_alignment

# Добавьте в импорты (если нет)
import os
from pathlib import Path

# Обновите конфигурацию Google Sheets (замените текущий блок)
GOOGLE_CREDS_PATH = "/app/credentials/service-account.json"
if not Path(GOOGLE_CREDS_PATH).exists():
    logger.error(f"Google credentials file not found at {GOOGLE_CREDS_PATH}")

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)

class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


# Стили для Google Sheets
HEADER_STYLE = {
    "backgroundColor": {"red": 0.20, "green": 0.39, "blue": 0.64},  # Темно-синий
    "textFormat": {"bold": True, "fontSize": 11, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
    "horizontalAlignment": "CENTER",
    "verticalAlignment": "MIDDLE",
    "wrapStrategy": "WRAP",
    "borders": {
        "top": {"style": "SOLID", "width": 1},
        "bottom": {"style": "SOLID", "width": 1},
        "left": {"style": "SOLID", "width": 1},
        "right": {"style": "SOLID", "width": 1}
    }
}

PRODUCT_ROW_STYLE = {
    "backgroundColor": {"red": 1, "green": 1, "blue": 1},  # Белый
    "textFormat": {"fontSize": 10},
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "wrapStrategy": "WRAP",
    "borders": {
        "top": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
        "bottom": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
        "left": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
        "right": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}}
    }
}

SUMMARY_ROW_STYLE = {
    "backgroundColor": {"red": 0.85, "green": 0.88, "blue": 0.95},  # Светло-синий
    "textFormat": {"bold": True, "fontSize": 10},
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "wrapStrategy": "WRAP",
    "borders": HEADER_STYLE["borders"]
}

TOTAL_ROW_STYLE = {
    "backgroundColor": {"red": 0.85, "green": 0.85, "blue": 0.85},  # Серый
    "textFormat": {"bold": True, "fontSize": 10},
    "horizontalAlignment": "RIGHT",
    "verticalAlignment": "MIDDLE",
    "borders": HEADER_STYLE["borders"]
}

NEGATIVE_PROFIT_STYLE = {
    "backgroundColor": {"red": 1, "green": 0.8, "blue": 0.8},  # Светло-красный
    "textFormat": {"bold": True}
}

NUMBER_FORMAT = {
    "numberFormat": {"type": "NUMBER", "pattern": "#,##0.00"},
    "horizontalAlignment": "RIGHT"
}

DATE_FORMAT = {
    "numberFormat": {"type": "DATE", "pattern": "dd.mm.yyyy"},
    "horizontalAlignment": "CENTER"
}

# Функция для подготовки значений
def prepare_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, Decimal):
        return float(value)
    elif value is None:
        return ""
    return value

@app.post("/api/export/gsheet")
async def export_to_gsheet(date_range: DateRange):
    try:
        logger.info("Создание Google Таблицы с оформлением как в Excel...")
        
        # Проверка учетных данных
        if not os.path.exists(GOOGLE_CREDS_PATH):
            logger.error("Файл учетных данных не найден!")
            return JSONResponse(
                status_code=500,
                content={"detail": "Файл учетных данных Google не найден"}
            )

        # Инициализация Google Sheets API
        gc = gspread.service_account(filename=GOOGLE_CREDS_PATH)
        
        # Создаем новую таблицу
        title = f"Отчет по отгрузкам {date_range.start_date} - {date_range.end_date}"
        sh = gc.create(title)
        sh.share(None, perm_type='anyone', role='writer')

        # Удаляем дефолтный лист
        if len(sh.worksheets()) > 1:
            sh.del_worksheet(sh.get_worksheet(0))

        # Стили оформления (добавляем стиль для сводного отчета)
        SUMMARY_HEADER_STYLE = {
            "backgroundColor": {"red": 0.31, "green": 0.60, "blue": 0.83},  # Более светлый синий
            "textFormat": {"bold": True, "fontSize": 11, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "wrapStrategy": "WRAP",
            "borders": HEADER_STYLE["borders"]
        }

        # ===== 1. ЛИСТ С ТОВАРАМИ =====
        worksheet_positions = sh.add_worksheet(title="Отчет по товарам", rows=1000, cols=33)
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Получаем данные с преобразованием типов
        cur.execute("""
            SELECT 
                d.number, d.date, d.counterparty, d.store, d.project, d.sales_channel,
                dp.product_name, dp.quantity, dp.price, dp.amount, 
                dp.cost_price, dp.article, dp.code, dp.overhead, dp.profit,
                d.promo_period, d.delivery_amount, d.admin_data,
                d.gdeslon, d.cityads, d.ozon, d.ozon_fbs,
                d.yamarket_fbs, d.yamarket_dbs, d.yandex_direct,
                d.price_ru, d.wildberries, d.gis2, d.seo,
                d.programmatic, d.avito, d.multiorders,
                d.estimated_discount
            FROM demand_positions dp
            JOIN demands d ON dp.demand_id = d.id
            WHERE d.date BETWEEN %s AND %s
            ORDER BY d.number, d.date DESC
        """, (date_range.start_date, date_range.end_date))
        
        # Преобразуем данные
        positions = []
        for row in cur.fetchall():
            positions.append([prepare_value(value) for value in row])
        
        # Заголовки
        pos_headers = [
            "Номер отгрузки", "Дата", "Контрагент", "Склад", "Проект", "Канал продаж",
            "Товар", "Количество", "Цена", "Сумма", "Себестоимость", "Артикул", "Код",
            "Накладные расходы", "Прибыль", "Акционный период", "Сумма доставки", "Адмидат",
            "ГдеСлон", "CityAds", "Ozon", "Ozon FBS", "Яндекс Маркет FBS", "Яндекс Маркет DBS",
            "Яндекс Директ", "Price ru", "Wildberries", "2Gis", "SEO", "Программатик", "Авито",
            "Мультиканальные заказы", "Примерная скидка"
        ]
        
        # Добавляем заголовки
        worksheet_positions.append_row(pos_headers)
        
        # Подготовка данных для вставки с группировкой
        if positions:
            current_demand = None
            rows_to_add = []
            batch_size = 100
            total_rows = 0
            
            for row in positions:
                demand_number = row[0]
                
                if demand_number != current_demand:
                    current_demand = demand_number
                    
                    # Получаем общую себестоимость по отгрузке
                    cur.execute("""
                        SELECT cost_price FROM demands 
                        WHERE number = %s AND date BETWEEN %s AND %s
                        LIMIT 1
                    """, (demand_number, date_range.start_date, date_range.end_date))
                    total_cost = prepare_value(cur.fetchone()[0]) if cur.rowcount > 0 else 0
                    
                    # Строка с итогами по отгрузке
                    summary_row = [
                        demand_number, row[1], row[2], row[3], row[4], row[5],
                        "Итого по отгрузке:", "", "", row[9], total_cost, "", "",
                        row[13], row[14], row[15], row[16], row[17], row[18],
                        row[19], row[20], row[21], row[22], row[23], row[24],
                        row[25], row[26], row[27], row[28], row[29], row[30],
                        row[31]
                    ]
                    rows_to_add.append([prepare_value(value) for value in summary_row])
                    total_rows += 1
                
                # Строка с товаром
                product_row = [
                    "", "", "", "", "", "",
                    row[6], row[7], row[8], row[9], row[10], row[11], row[12],
                    "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""
                ]
                rows_to_add.append([prepare_value(value) for value in product_row])
                total_rows += 1
            
            # Вставляем данные пакетами
            for i in range(0, len(rows_to_add), batch_size):
                batch = rows_to_add[i:i + batch_size]
                worksheet_positions.append_rows(batch)
        
        # Форматируем лист с товарами
        last_row = total_rows + 1 if positions else 1
        requests = []
        
        # Форматирование заголовков
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_positions.id,
                    "startRowIndex": 0,
                    "endRowIndex": 1
                },
                "cell": {"userEnteredFormat": HEADER_STYLE},
                "fields": "userEnteredFormat"
            }
        })
        
        # Форматирование строк с товарами
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_positions.id,
                    "startRowIndex": 1,
                    "endRowIndex": last_row
                },
                "cell": {"userEnteredFormat": PRODUCT_ROW_STYLE},
                "fields": "userEnteredFormat"
            }
        })
        
        # Форматирование строк с итогами по отгрузке
        if positions:
            for i, row in enumerate(rows_to_add, start=1):
                if row[6] == "Итого по отгрузке:":
                    requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": worksheet_positions.id,
                                "startRowIndex": i,
                                "endRowIndex": i + 1
                            },
                            "cell": {"userEnteredFormat": SUMMARY_ROW_STYLE},
                            "fields": "userEnteredFormat"
                        }
                    })
                    
                    # Выравнивание "Итого по отгрузке:" по правому краю
                    requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": worksheet_positions.id,
                                "startRowIndex": i,
                                "endRowIndex": i + 1,
                                "startColumnIndex": 6,
                                "endColumnIndex": 7
                            },
                            "cell": {"userEnteredFormat": {"horizontalAlignment": "RIGHT"}},
                            "fields": "userEnteredFormat.horizontalAlignment"
                        }
                    })
        
        # Форматирование числовых столбцов
        numeric_columns = [7, 8, 9, 10, 13, 14] + list(range(16, 32))
        for col in numeric_columns:
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_positions.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_row,
                        "startColumnIndex": col,
                        "endColumnIndex": col + 1
                    },
                    "cell": {"userEnteredFormat": NUMBER_FORMAT},
                    "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
                }
            })
        
        # Форматирование дат
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_positions.id,
                    "startRowIndex": 1,
                    "endRowIndex": last_row,
                    "startColumnIndex": 1,
                    "endColumnIndex": 2
                },
                "cell": {"userEnteredFormat": DATE_FORMAT},
                "fields": "userEnteredFormat"
            }
        })
        
        # Подсветка отрицательной прибыли
        requests.append({
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{
                        "sheetId": worksheet_positions.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_row,
                        "startColumnIndex": 14,
                        "endColumnIndex": 15
                    }],
                    "booleanRule": {
                        "condition": {
                            "type": "NUMBER_LESS",
                            "values": [{"userEnteredValue": "0"}]
                        },
                        "format": NEGATIVE_PROFIT_STYLE
                    }
                },
                "index": 0
            }
        })
        
        # Установка ширины столбцов
        column_widths = [
            {"pixelSize": 100},  # A: Номер отгрузки
            {"pixelSize": 150},  # B: Дата
            {"pixelSize": 200},  # C: Контрагент
            {"pixelSize": 120},  # D: Склад
            {"pixelSize": 120},  # E: Проект
            {"pixelSize": 150},  # F: Канал продаж
            {"pixelSize": 300},  # G: Товар
            {"pixelSize": 90},   # H: Количество
            {"pixelSize": 90},   # I: Цена
            {"pixelSize": 90},   # J: Сумма
            {"pixelSize": 110},  # K: Себестоимость
            {"pixelSize": 100},  # L: Артикул
            {"pixelSize": 80},   # M: Код
            {"pixelSize": 110},  # N: Накладные расходы
            {"pixelSize": 90},   # O: Прибыль
            {"pixelSize": 120},  # P: Акционный период
            {"pixelSize": 110},  # Q: Сумма доставки
            {"pixelSize": 90},   # R: Адмидат
            {"pixelSize": 90},   # S: ГдеСлон
            {"pixelSize": 90},   # T: CityAds
            {"pixelSize": 80},   # U: Ozon
            {"pixelSize": 100},  # V: Ozon FBS
            {"pixelSize": 130},  # W: Яндекс Маркет FBS
            {"pixelSize": 130},  # X: Яндекс Маркет DBS
            {"pixelSize": 110},  # Y: Яндекс Директ
            {"pixelSize": 90},   # Z: Price ru
            {"pixelSize": 110},  # AA: Wildberries
            {"pixelSize": 80},   # AB: 2Gis
            {"pixelSize": 80},   # AC: SEO
            {"pixelSize": 110},  # AD: Программатик
            {"pixelSize": 80},   # AE: Авито
            {"pixelSize": 140},  # AF: Мультиканальные заказы
            {"pixelSize": 120}   # AG: Примерная скидка
        ]
        
        for i, width in enumerate(column_widths):
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet_positions.id,
                        "dimension": "COLUMNS",
                        "startIndex": i,
                        "endIndex": i + 1
                    },
                    "properties": width,
                    "fields": "pixelSize"
                }
            })
        
        # Добавляем итоговую строку
        if positions:
            # Формулы для суммирования
            sum_formulas = [
                "", "", "", "", "", "", "Общий итог:",
                f'=SUM(H2:H{last_row})',
                f'=AVERAGE(I2:I{last_row})',
                f'=SUM(J2:J{last_row})',
                f'=SUM(K2:K{last_row})',
                "", "",
                f'=SUM(N2:N{last_row})',
                f'=SUM(O2:O{last_row})',
                "",
                f'=SUM(Q2:Q{last_row})',
                f'=SUM(R2:R{last_row})',
                f'=SUM(S2:S{last_row})',
                f'=SUM(T2:T{last_row})',
                f'=SUM(U2:U{last_row})',
                f'=SUM(V2:V{last_row})',
                f'=SUM(W2:W{last_row})',
                f'=SUM(X2:X{last_row})',
                f'=SUM(Y2:Y{last_row})',
                f'=SUM(Z2:Z{last_row})',
                f'=SUM(AA2:AA{last_row})',
                f'=SUM(AB2:AB{last_row})',
                f'=SUM(AC2:AC{last_row})',
                f'=SUM(AD2:AD{last_row})',
                f'=SUM(AE2:AE{last_row})',
                f'=SUM(AF2:AF{last_row})',
                f'=SUM(AG2:AG{last_row})'
            ]
            
            worksheet_positions.append_row(sum_formulas)
            
            # Форматирование итоговой строки
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_positions.id,
                        "startRowIndex": last_row,
                        "endRowIndex": last_row + 1
                    },
                    "cell": {"userEnteredFormat": TOTAL_ROW_STYLE},
                    "fields": "userEnteredFormat"
                }
            })
            
            last_row += 1
        
        # Фильтры и закрепление
        requests.extend([
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": worksheet_positions.id,
                            "startRowIndex": 0,
                            "endRowIndex": last_row,
                            "startColumnIndex": 0,
                            "endColumnIndex": 33
                        }
                    }
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": worksheet_positions.id,
                        "gridProperties": {"frozenRowCount": 1}
                    },
                    "fields": "gridProperties.frozenRowCount"
                }
            }
        ])
        
        # ===== 2. ЛИСТ С ОТГРУЗКАМИ =====
        worksheet_demands = sh.add_worksheet(title="Отчет по отгрузкам", rows=1000, cols=28)
        
        # Получаем данные с преобразованием типов
        cur.execute("""
            SELECT 
                number, date, counterparty, store, project, sales_channel,
                amount, cost_price, overhead, profit, 
                promo_period, delivery_amount, admin_data,
                gdeslon, cityads, ozon, ozon_fbs,
                yamarket_fbs, yamarket_dbs, yandex_direct,
                price_ru, wildberries, gis2, seo,
                programmatic, avito, multiorders,
                estimated_discount
            FROM demands
            WHERE date BETWEEN %s AND %s
            ORDER BY date DESC
        """, (date_range.start_date, date_range.end_date))
        
        # Преобразуем данные
        demands = []
        for row in cur.fetchall():
            demands.append([prepare_value(value) for value in row])
        
        conn.close()
        
        # Заголовки
        demands_headers = [
            "Номер отгрузки", "Дата", "Контрагент", "Склад", "Проект", "Канал продаж",
            "Сумма", "Себестоимость", "Накладные расходы", "Прибыль", "Акционный период",
            "Сумма доставки", "Адмидат", "ГдеСлон", "CityAds", "Ozon", "Ozon FBS",
            "Яндекс Маркет FBS", "Яндекс Маркет DBS", "Яндекс Директ", "Price ru",
            "Wildberries", "2Gis", "SEO", "Программатик", "Авито", "Мультиканальные заказы",
            "Примерная скидка"
        ]
        
        # Добавляем заголовки и данные
        worksheet_demands.append_row(demands_headers)
        if demands:
            worksheet_demands.append_rows(demands)
        
        # Форматируем лист с отгрузками
        last_demand_row = len(demands) + 1 if demands else 1
        demand_requests = []
        
        # Форматирование заголовков
        demand_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_demands.id,
                    "startRowIndex": 0,
                    "endRowIndex": 1
                },
                "cell": {"userEnteredFormat": HEADER_STYLE},
                "fields": "userEnteredFormat"
            }
        })
        
        # Форматирование данных
        demand_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_demands.id,
                    "startRowIndex": 1,
                    "endRowIndex": last_demand_row
                },
                "cell": {"userEnteredFormat": PRODUCT_ROW_STYLE},
                "fields": "userEnteredFormat"
            }
        })
        
        # Форматирование числовых столбцов
        numeric_demand_columns = [6, 7, 8, 9, 11] + list(range(12, 28))
        for col in numeric_demand_columns:
            demand_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_demands.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_demand_row,
                        "startColumnIndex": col,
                        "endColumnIndex": col + 1
                    },
                    "cell": {"userEnteredFormat": NUMBER_FORMAT},
                    "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
                }
            })
        
        # Форматирование дат
        demand_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_demands.id,
                    "startRowIndex": 1,
                    "endRowIndex": last_demand_row,
                    "startColumnIndex": 1,
                    "endColumnIndex": 2
                },
                "cell": {"userEnteredFormat": DATE_FORMAT},
                "fields": "userEnteredFormat"
            }
        })
        
        # Подсветка отрицательной прибыли
        demand_requests.append({
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{
                        "sheetId": worksheet_demands.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_demand_row,
                        "startColumnIndex": 9,
                        "endColumnIndex": 10
                    }],
                    "booleanRule": {
                        "condition": {
                            "type": "NUMBER_LESS",
                            "values": [{"userEnteredValue": "0"}]
                        },
                        "format": NEGATIVE_PROFIT_STYLE
                    }
                },
                "index": 0
            }
        })
        
        # Добавляем итоговую строку
        if demands:
            # Формулы для суммирования
            sum_formulas = [
                "Итого:", "", "", "", "", "",
                f'=SUM(G2:G{last_demand_row})',
                f'=SUM(H2:H{last_demand_row})',
                f'=SUM(I2:I{last_demand_row})',
                f'=SUM(J2:J{last_demand_row})',
                "",
                f'=SUM(L2:L{last_demand_row})',
                f'=SUM(M2:M{last_demand_row})',
                f'=SUM(N2:N{last_demand_row})',
                f'=SUM(O2:O{last_demand_row})',
                f'=SUM(P2:P{last_demand_row})',
                f'=SUM(Q2:Q{last_demand_row})',
                f'=SUM(R2:R{last_demand_row})',
                f'=SUM(S2:S{last_demand_row})',
                f'=SUM(T2:T{last_demand_row})',
                f'=SUM(U2:U{last_demand_row})',
                f'=SUM(V2:V{last_demand_row})',
                f'=SUM(W2:W{last_demand_row})',
                f'=SUM(X2:X{last_demand_row})',
                f'=SUM(Y2:Y{last_demand_row})',
                f'=SUM(Z2:Z{last_demand_row})',
                f'=SUM(AA2:AA{last_demand_row})',
                f'=SUM(AB2:AB{last_demand_row})'
            ]
            
            worksheet_demands.append_row(sum_formulas)
            
            # Форматирование итоговой строки
            demand_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_demands.id,
                        "startRowIndex": last_demand_row,
                        "endRowIndex": last_demand_row + 1
                    },
                    "cell": {"userEnteredFormat": TOTAL_ROW_STYLE},
                    "fields": "userEnteredFormat"
                }
            })
            
            last_demand_row += 1
        
        # Установка ширины столбцов
        demand_column_widths = [
            {"pixelSize": 100},  # A: Номер отгрузки
            {"pixelSize": 150},  # B: Дата
            {"pixelSize": 200},  # C: Контрагент
            {"pixelSize": 120},  # D: Склад
            {"pixelSize": 120},  # E: Проект
            {"pixelSize": 150},  # F: Канал продаж
            {"pixelSize": 90},   # G: Сумма
            {"pixelSize": 110},  # H: Себестоимость
            {"pixelSize": 110},  # I: Накладные расходы
            {"pixelSize": 90},   # J: Прибыль
            {"pixelSize": 120},  # K: Акционный период
            {"pixelSize": 110},  # L: Сумма доставки
            {"pixelSize": 90},   # M: Адмидат
            {"pixelSize": 90},   # N: ГдеСлон
            {"pixelSize": 90},   # O: CityAds
            {"pixelSize": 80},   # P: Ozon
            {"pixelSize": 100},  # Q: Ozon FBS
            {"pixelSize": 130},  # R: Яндекс Маркет FBS
            {"pixelSize": 130},  # S: Яндекс Маркет DBS
            {"pixelSize": 110},  # T: Яндекс Директ
            {"pixelSize": 90},   # U: Price ru
            {"pixelSize": 110},  # V: Wildberries
            {"pixelSize": 80},   # W: 2Gis
            {"pixelSize": 80},   # X: SEO
            {"pixelSize": 110},  # Y: Программатик
            {"pixelSize": 80},   # Z: Авито
            {"pixelSize": 140},  # AA: Мультиканальные заказы
            {"pixelSize": 120}   # AB: Примерная скидка
        ]
        
        for i, width in enumerate(demand_column_widths):
            demand_requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet_demands.id,
                        "dimension": "COLUMNS",
                        "startIndex": i,
                        "endIndex": i + 1
                    },
                    "properties": width,
                    "fields": "pixelSize"
                }
            })
        
        # Фильтры и закрепление
        demand_requests.extend([
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": worksheet_demands.id,
                            "startRowIndex": 0,
                            "endRowIndex": last_demand_row,
                            "startColumnIndex": 0,
                            "endColumnIndex": 28
                        }
                    }
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": worksheet_demands.id,
                        "gridProperties": {"frozenRowCount": 1}
                    },
                    "fields": "gridProperties.frozenRowCount"
                }
            }
        ])
        
        # ===== 3. ЛИСТ СО СВОДНЫМ ОТЧЕТОМ ПО ТОВАРАМ =====
        worksheet_summary = sh.add_worksheet(title="Сводный отчет по товарам", rows=1000, cols=14)
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Получаем данные для сводного отчета
        cur.execute("""
            SELECT 
                dp.product_name,
                dp.article,
                dp.code,
                SUM(dp.quantity) as total_quantity,
                d.store,
                d.project,
                d.sales_channel,
                AVG(dp.price) as avg_price,
                SUM(d.delivery_amount * (dp.amount / NULLIF(d.amount, 0))) as delivery_share,
                SUM(dp.amount) as total_amount,
                SUM(dp.cost_price) as total_cost_price,
                SUM(d.overhead * (dp.amount / NULLIF(d.amount, 0))) as overhead_share,
                SUM(dp.amount - dp.cost_price - (d.overhead * (dp.amount / NULLIF(d.amount, 0)))) as total_profit,
                CASE 
                    WHEN SUM(dp.amount) = 0 THEN 0 
                    ELSE (SUM(dp.amount) - SUM(dp.cost_price) - SUM(d.overhead * (dp.amount / NULLIF(d.amount, 0)))) / SUM(dp.amount) 
                END as margin
            FROM demand_positions dp
            JOIN demands d ON dp.demand_id = d.id
            WHERE d.date BETWEEN %s AND %s
            GROUP BY dp.product_name, dp.article, dp.code, d.store, d.project, d.sales_channel
            ORDER BY total_amount DESC
        """, (date_range.start_date, date_range.end_date))
        
        # Преобразуем данные
        summary_data = []
        for row in cur.fetchall():
            summary_data.append([prepare_value(value) for value in row])
        
        conn.close()
        
        # Заголовки
        summary_headers = [
            "Товар", "Артикул", "Код", "Общее количество", "Склад", "Проект", 
            "Канал продаж", "Средняя цена", "Сумма оплачиваемой доставки", 
            "Общая сумма", "Себестоимость товара", "Сумма накладных расходов", 
            "Общая прибыль", "Маржинальность"
        ]
        
        # Добавляем заголовки
        worksheet_summary.append_row(summary_headers)
        
        # Добавляем данные
        if summary_data:
            worksheet_summary.append_rows(summary_data)
        
        # Форматируем лист со сводным отчетом
        last_summary_row = len(summary_data) + 1 if summary_data else 1
        summary_requests = []
        
        # Форматирование заголовков
        summary_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_summary.id,
                    "startRowIndex": 0,
                    "endRowIndex": 1
                },
                "cell": {"userEnteredFormat": SUMMARY_HEADER_STYLE},
                "fields": "userEnteredFormat"
            }
        })
        
        # Форматирование данных
        summary_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_summary.id,
                    "startRowIndex": 1,
                    "endRowIndex": last_summary_row
                },
                "cell": {"userEnteredFormat": PRODUCT_ROW_STYLE},
                "fields": "userEnteredFormat"
            }
        })
        
        # Форматирование числовых столбцов
        numeric_summary_columns = [3, 7, 8, 9, 10, 11, 12]  # Индексы столбцов с числами
        percent_column = 13  # Маржинальность
        
        for col in numeric_summary_columns:
            summary_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_summary.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_summary_row,
                        "startColumnIndex": col,
                        "endColumnIndex": col + 1
                    },
                    "cell": {"userEnteredFormat": NUMBER_FORMAT},
                    "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
                }
            })
        
        # Форматирование столбца с маржинальностью
        summary_requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet_summary.id,
                    "startRowIndex": 1,
                    "endRowIndex": last_summary_row,
                    "startColumnIndex": percent_column,
                    "endColumnIndex": percent_column + 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "PERCENT", "pattern": "0.00%"},
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        })
        
        # Подсветка отрицательной маржинальности
        summary_requests.append({
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{
                        "sheetId": worksheet_summary.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_summary_row,
                        "startColumnIndex": percent_column,
                        "endColumnIndex": percent_column + 1
                    }],
                    "booleanRule": {
                        "condition": {
                            "type": "NUMBER_LESS",
                            "values": [{"userEnteredValue": "0"}]
                        },
                        "format": NEGATIVE_PROFIT_STYLE
                    }
                },
                "index": 0
            }
        })
        
        # Установка ширины столбцов
        summary_column_widths = [
            {"pixelSize": 250},  # A: Товар
            {"pixelSize": 100},  # B: Артикул
            {"pixelSize": 80},   # C: Код
            {"pixelSize": 110},  # D: Общее количество
            {"pixelSize": 120},  # E: Склад
            {"pixelSize": 120},  # F: Проект
            {"pixelSize": 150},  # G: Канал продаж
            {"pixelSize": 90},   # H: Средняя цена
            {"pixelSize": 140},  # I: Сумма оплачиваемой доставки
            {"pixelSize": 90},   # J: Общая сумма
            {"pixelSize": 120},  # K: Себестоимость товара
            {"pixelSize": 140},  # L: Сумма накладных расходов
            {"pixelSize": 100},  # M: Общая прибыль
            {"pixelSize": 100}   # N: Маржинальность
        ]
        
        for i, width in enumerate(summary_column_widths):
            summary_requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet_summary.id,
                        "dimension": "COLUMNS",
                        "startIndex": i,
                        "endIndex": i + 1
                    },
                    "properties": width,
                    "fields": "pixelSize"
                }
            })
        
        # Добавляем итоговую строку
        if summary_data:
            # Формулы для суммирования
            sum_formulas = [
                "Общий итог:", "", "", 
                f'=SUM(D2:D{last_summary_row})',
                "", "", "",
                f'=AVERAGE(H2:H{last_summary_row})',
                f'=SUM(I2:I{last_summary_row})',
                f'=SUM(J2:J{last_summary_row})',
                f'=SUM(K2:K{last_summary_row})',
                f'=SUM(L2:L{last_summary_row})',
                f'=SUM(M2:M{last_summary_row})',
                f'=IF(SUM(J2:J{last_summary_row})=0, 0, (SUM(J2:J{last_summary_row})-SUM(K2:K{last_summary_row})-SUM(L2:L{last_summary_row}))/SUM(J2:J{last_summary_row}))'
            ]
            
            worksheet_summary.append_row(sum_formulas)
            
            # Форматирование итоговой строки
            summary_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet_summary.id,
                        "startRowIndex": last_summary_row,
                        "endRowIndex": last_summary_row + 1
                    },
                    "cell": {"userEnteredFormat": TOTAL_ROW_STYLE},
                    "fields": "userEnteredFormat"
                }
            })
            
            last_summary_row += 1
        
        # Фильтры и закрепление
        summary_requests.extend([
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": worksheet_summary.id,
                            "startRowIndex": 0,
                            "endRowIndex": last_summary_row,
                            "startColumnIndex": 0,
                            "endColumnIndex": 14
                        }
                    }
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": worksheet_summary.id,
                        "gridProperties": {"frozenRowCount": 1}
                    },
                    "fields": "gridProperties.frozenRowCount"
                }
            }
        ])
        
        # Объединяем все запросы
        all_requests = requests + demand_requests + summary_requests
        
        # Применяем все запросы
        sh.batch_update({"requests": all_requests})
        
        # Устанавливаем порядок листов (товары первыми, затем отгрузки, затем сводный отчет)
        sh.reorder_worksheets([worksheet_positions, worksheet_demands, worksheet_summary])
        
        logger.info(f"Таблица создана с оформлением как в Excel: {sh.url}")
        return {"url": sh.url}
        
    except Exception as e:
        logger.error(f"Ошибка при создании таблицы: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"detail": f"Ошибка при создании таблицы: {str(e)}"}
        )